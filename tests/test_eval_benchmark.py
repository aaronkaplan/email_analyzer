"""Tests for the eval_benchmark module."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from email_analyzer.config import EvalBenchmarkConfig
from email_analyzer.eval_benchmark import (
    DEFAULT_CATEGORY_MAP,
    ClassMetrics,
    _compute_report,
    _infer_ground_truth,
    _load_predictions,
    PredictionRow,
    run_eval_benchmark,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_batch_output_record(
    custom_id: str,
    classification: str,
    confidence: float = 0.9,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Build one line of batch_output.jsonl matching the OpenAI-shaped format."""
    payload: dict[str, Any] = {
        "classification": classification,
        "confidence": confidence,
    }
    if extra:
        payload.update(extra)

    return {
        "custom_id": custom_id,
        "error": None,
        "response": {
            "status_code": 200,
            "body": {
                "output": [
                    {
                        "type": "message",
                        "content": [
                            {
                                "type": "output_text",
                                "text": json.dumps(payload, ensure_ascii=False),
                            }
                        ],
                    }
                ]
            },
        },
    }


def _write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    path.write_text(
        "\n".join(json.dumps(r, ensure_ascii=False, sort_keys=True) for r in records)
        + "\n",
        encoding="utf-8",
    )


# ---------------------------------------------------------------------------
# Ground truth inference
# ---------------------------------------------------------------------------


class TestInferGroundTruth:
    """Tests for _infer_ground_truth."""

    def test_standard_spamassassin_categories(self) -> None:
        assert (
            _infer_ground_truth("easy_ham__00001.abc123", DEFAULT_CATEGORY_MAP) == "ham"
        )
        assert (
            _infer_ground_truth("easy_ham_2__00264.def456", DEFAULT_CATEGORY_MAP)
            == "ham"
        )
        assert (
            _infer_ground_truth("hard_ham__00003.ghi789", DEFAULT_CATEGORY_MAP) == "ham"
        )
        assert _infer_ground_truth("spam__00010.jkl012", DEFAULT_CATEGORY_MAP) == "spam"
        assert (
            _infer_ground_truth("spam_2__00500.mno345", DEFAULT_CATEGORY_MAP) == "spam"
        )

    def test_custom_category_map(self) -> None:
        custom = {"phishing": "phishing", "legitimate": "legitimate"}
        assert _infer_ground_truth("phishing__msg001.eml", custom) == "phishing"
        assert _infer_ground_truth("legitimate__msg002.eml", custom) == "legitimate"

    def test_unknown_prefix_returned_as_is(self) -> None:
        assert _infer_ground_truth("unknown__001.eml", {}) == "unknown"

    def test_no_double_underscore_uses_full_id(self) -> None:
        assert _infer_ground_truth("spam", DEFAULT_CATEGORY_MAP) == "spam"

    def test_case_insensitive(self) -> None:
        assert _infer_ground_truth("EASY_HAM__001.eml", DEFAULT_CATEGORY_MAP) == "ham"
        assert _infer_ground_truth("Spam__001.eml", DEFAULT_CATEGORY_MAP) == "spam"


# ---------------------------------------------------------------------------
# ClassMetrics
# ---------------------------------------------------------------------------


class TestClassMetrics:
    """Tests for ClassMetrics dataclass properties."""

    def test_perfect_class(self) -> None:
        m = ClassMetrics(
            label="spam", true_positives=10, false_positives=0, false_negatives=0
        )
        assert m.precision == 1.0
        assert m.recall == 1.0
        assert m.f1 == 1.0
        assert m.support == 10

    def test_zero_support(self) -> None:
        m = ClassMetrics(
            label="empty", true_positives=0, false_positives=0, false_negatives=0
        )
        assert m.precision == 0.0
        assert m.recall == 0.0
        assert m.f1 == 0.0
        assert m.support == 0

    def test_mixed(self) -> None:
        m = ClassMetrics(
            label="spam", true_positives=8, false_positives=2, false_negatives=3
        )
        assert m.precision == pytest.approx(0.8)
        assert m.recall == pytest.approx(8 / 11)
        assert m.support == 11

    def test_precision_only_fp(self) -> None:
        m = ClassMetrics(
            label="x", true_positives=0, false_positives=5, false_negatives=0
        )
        assert m.precision == 0.0
        assert m.recall == 0.0


# ---------------------------------------------------------------------------
# Load predictions
# ---------------------------------------------------------------------------


class TestLoadPredictions:
    """Tests for _load_predictions."""

    def test_basic_loading(self, tmp_path: Path) -> None:
        records = [
            _make_batch_output_record("spam__001.abc", "spam", 0.95),
            _make_batch_output_record("easy_ham__001.def", "ham", 0.8),
        ]
        jsonl = tmp_path / "batch_output.jsonl"
        _write_jsonl(jsonl, records)

        rows = _load_predictions(jsonl, "classification", DEFAULT_CATEGORY_MAP)
        assert len(rows) == 2
        assert rows[0].custom_id == "spam__001.abc"
        assert rows[0].ground_truth == "spam"
        assert rows[0].predicted == "spam"
        assert rows[0].confidence == pytest.approx(0.95)
        assert rows[1].ground_truth == "ham"
        assert rows[1].predicted == "ham"

    def test_skips_error_records(self, tmp_path: Path) -> None:
        records = [
            {
                "custom_id": "spam__001.abc",
                "error": {"message": "timeout"},
                "response": None,
            },
            _make_batch_output_record("easy_ham__001.def", "ham"),
        ]
        jsonl = tmp_path / "batch_output.jsonl"
        _write_jsonl(jsonl, records)

        rows = _load_predictions(jsonl, "classification", DEFAULT_CATEGORY_MAP)
        assert len(rows) == 1
        assert rows[0].custom_id == "easy_ham__001.def"

    def test_skips_blank_lines(self, tmp_path: Path) -> None:
        jsonl = tmp_path / "batch_output.jsonl"
        record = _make_batch_output_record("spam__001.abc", "spam")
        jsonl.write_text(
            "\n" + json.dumps(record) + "\n\n",
            encoding="utf-8",
        )

        rows = _load_predictions(jsonl, "classification", DEFAULT_CATEGORY_MAP)
        assert len(rows) == 1

    def test_custom_label_field(self, tmp_path: Path) -> None:
        payload = {"category": "spam", "confidence": 0.7}
        record: dict[str, Any] = {
            "custom_id": "spam__001.abc",
            "error": None,
            "response": {
                "status_code": 200,
                "body": {
                    "output": [
                        {
                            "type": "message",
                            "content": [
                                {
                                    "type": "output_text",
                                    "text": json.dumps(payload),
                                }
                            ],
                        }
                    ]
                },
            },
        }
        jsonl = tmp_path / "batch_output.jsonl"
        _write_jsonl(jsonl, [record])

        rows = _load_predictions(jsonl, "category", DEFAULT_CATEGORY_MAP)
        assert rows[0].predicted == "spam"


# ---------------------------------------------------------------------------
# Compute report
# ---------------------------------------------------------------------------


class TestComputeReport:
    """Tests for _compute_report."""

    def test_perfect_binary_classification(self) -> None:
        rows = [
            PredictionRow("spam__1", "spam", "spam"),
            PredictionRow("spam__2", "spam", "spam"),
            PredictionRow("ham__1", "ham", "ham"),
            PredictionRow("ham__2", "ham", "ham"),
        ]
        report = _compute_report(rows)

        assert report.total == 4
        assert report.correct == 4
        assert report.accuracy == 1.0
        assert report.macro_f1 == 1.0
        assert report.weighted_f1 == 1.0

        assert report.confusion == {
            "ham": {"ham": 2, "spam": 0},
            "spam": {"ham": 0, "spam": 2},
        }
        assert report.per_class["spam"].precision == 1.0
        assert report.per_class["ham"].recall == 1.0

    def test_imperfect_classification(self) -> None:
        rows = [
            PredictionRow("spam__1", "spam", "spam"),
            PredictionRow("spam__2", "spam", "ham"),  # false negative
            PredictionRow("ham__1", "ham", "ham"),
            PredictionRow("ham__2", "ham", "spam"),  # false positive
        ]
        report = _compute_report(rows)

        assert report.total == 4
        assert report.correct == 2
        assert report.accuracy == 0.5

        # Spam: TP=1, FP=1, FN=1 -> P=0.5, R=0.5, F1=0.5
        assert report.per_class["spam"].precision == pytest.approx(0.5)
        assert report.per_class["spam"].recall == pytest.approx(0.5)
        assert report.per_class["spam"].f1 == pytest.approx(0.5)

    def test_multiclass(self) -> None:
        rows = [
            PredictionRow("a__1", "a", "a"),
            PredictionRow("b__1", "b", "b"),
            PredictionRow("c__1", "c", "a"),  # c misclassified as a
        ]
        report = _compute_report(rows)

        assert report.total == 3
        assert report.correct == 2
        assert set(report.labels) == {"a", "b", "c"}

        # a: TP=1, FP=1 (the c->a), FN=0 -> P=0.5, R=1.0
        assert report.per_class["a"].precision == pytest.approx(0.5)
        assert report.per_class["a"].recall == pytest.approx(1.0)

        # c: TP=0, FP=0, FN=1 -> P=0, R=0, F1=0
        assert report.per_class["c"].f1 == 0.0

    def test_empty_rows(self) -> None:
        report = _compute_report([])
        assert report.total == 0
        assert report.accuracy == 0.0
        assert report.labels == []


# ---------------------------------------------------------------------------
# End-to-end: run_eval_benchmark
# ---------------------------------------------------------------------------


class TestRunEvalBenchmark:
    """Integration tests for the full run_eval_benchmark entry point."""

    def test_basic_spam_ham_evaluation(self, tmp_path: Path) -> None:
        records = [
            _make_batch_output_record("spam__001.abc", "spam", 0.95),
            _make_batch_output_record("spam__002.def", "spam", 0.85),
            _make_batch_output_record("spam_2__001.ghi", "spam", 0.90),
            _make_batch_output_record("easy_ham__001.jkl", "ham", 0.88),
            _make_batch_output_record("easy_ham_2__001.mno", "ham", 0.92),
            _make_batch_output_record("hard_ham__001.pqr", "ham", 0.70),
            # One misclassification: ham predicted as spam
            _make_batch_output_record("easy_ham__002.stu", "spam", 0.55),
        ]

        jsonl = tmp_path / "batch_output.jsonl"
        _write_jsonl(jsonl, records)
        output_xlsx = tmp_path / "eval_report.xlsx"

        exit_code = run_eval_benchmark(
            EvalBenchmarkConfig(
                batch_output_jsonl=jsonl,
                output_xlsx=output_xlsx,
            )
        )

        assert exit_code == 0
        assert output_xlsx.exists()
        assert output_xlsx.with_suffix(".json").exists()

        # Verify XLSX structure.
        wb = load_workbook(output_xlsx)
        assert "Summary" in wb.sheetnames
        assert "Confusion Matrix" in wb.sheetnames
        assert "Raw Data" in wb.sheetnames

        # Verify JSON report.
        report_data = json.loads(output_xlsx.with_suffix(".json").read_text())
        assert report_data["total"] == 7
        assert report_data["correct"] == 6
        assert report_data["accuracy"] == pytest.approx(6 / 7)
        assert set(report_data["labels"]) == {"ham", "spam"}

    def test_custom_category_map(self, tmp_path: Path) -> None:
        category_map = {"phishing": "phishing", "legit": "legitimate"}
        map_file = tmp_path / "cat_map.json"
        map_file.write_text(json.dumps(category_map), encoding="utf-8")

        records = [
            _make_batch_output_record("phishing__001.abc", "phishing", 0.9),
            _make_batch_output_record("legit__001.def", "legitimate", 0.85),
        ]
        jsonl = tmp_path / "batch_output.jsonl"
        _write_jsonl(jsonl, records)
        output_xlsx = tmp_path / "eval.xlsx"

        exit_code = run_eval_benchmark(
            EvalBenchmarkConfig(
                batch_output_jsonl=jsonl,
                output_xlsx=output_xlsx,
                category_map_file=map_file,
            )
        )

        assert exit_code == 0
        report_data = json.loads(output_xlsx.with_suffix(".json").read_text())
        assert report_data["accuracy"] == 1.0
        assert set(report_data["labels"]) == {"legitimate", "phishing"}

    def test_default_output_xlsx_path(self, tmp_path: Path) -> None:
        records = [_make_batch_output_record("spam__001.abc", "spam")]
        jsonl = tmp_path / "batch_output.jsonl"
        _write_jsonl(jsonl, records)

        exit_code = run_eval_benchmark(EvalBenchmarkConfig(batch_output_jsonl=jsonl))
        assert exit_code == 0

        default_xlsx = jsonl.with_suffix(".eval.xlsx")
        assert default_xlsx.exists()

    def test_no_valid_rows_returns_1(self, tmp_path: Path) -> None:
        jsonl = tmp_path / "batch_output.jsonl"
        jsonl.write_text("", encoding="utf-8")

        exit_code = run_eval_benchmark(EvalBenchmarkConfig(batch_output_jsonl=jsonl))
        assert exit_code == 1

    def test_xlsx_raw_data_sheet_row_count(self, tmp_path: Path) -> None:
        records = [
            _make_batch_output_record(
                f"spam__{i:03d}.abc",
                "spam",
                extra={"reason": f"reason {i}", "indicators": [f"ind_{i}"]},
            )
            for i in range(5)
        ]
        records.extend(
            _make_batch_output_record(
                f"easy_ham__{i:03d}.def",
                "ham",
                extra={"reason": f"ok {i}", "indicators": []},
            )
            for i in range(3)
        )
        jsonl = tmp_path / "batch_output.jsonl"
        _write_jsonl(jsonl, records)
        output_xlsx = tmp_path / "eval.xlsx"

        run_eval_benchmark(
            EvalBenchmarkConfig(batch_output_jsonl=jsonl, output_xlsx=output_xlsx)
        )

        wb = load_workbook(output_xlsx)
        ws_raw = wb["Raw Data"]
        # Header row + 8 data rows.
        assert ws_raw.max_row == 9

    def test_confusion_matrix_sheet_structure(self, tmp_path: Path) -> None:
        records = [
            _make_batch_output_record("spam__001.abc", "spam"),
            _make_batch_output_record("easy_ham__001.def", "ham"),
            _make_batch_output_record("easy_ham__002.ghi", "spam"),
        ]
        jsonl = tmp_path / "batch_output.jsonl"
        _write_jsonl(jsonl, records)
        output_xlsx = tmp_path / "eval.xlsx"

        run_eval_benchmark(
            EvalBenchmarkConfig(batch_output_jsonl=jsonl, output_xlsx=output_xlsx)
        )

        wb = load_workbook(output_xlsx)
        ws_cm = wb["Confusion Matrix"]

        # Header row: empty cell + labels.
        header = [cell.value for cell in ws_cm[1]]
        assert header[0] is None or header[0] == ""
        assert "ham" in header
        assert "spam" in header
