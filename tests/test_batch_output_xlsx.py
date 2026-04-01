from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from email_analyzer.batch_output_xlsx import run_batch_output_to_xlsx
from email_analyzer.config import BatchOutputXlsxConfig


def test_batch_output_to_xlsx_writes_filename_and_schema_columns(
    tmp_path: Path,
) -> None:
    input_jsonl = tmp_path / "batch_output.jsonl"
    output_xlsx = tmp_path / "batch_output.xlsx"
    schema_file = tmp_path / "schema.py"

    schema_file.write_text(
        "from pydantic import BaseModel, Field\n\n"
        "class mySchema(BaseModel):\n"
        '    sender: str = Field(alias="from")\n'
        "    recipient: str\n"
        "    action_items: list[str] = Field(default_factory=list)\n"
        "    is_spam: bool\n",
        encoding="utf-8",
    )

    output_payload = {
        "from": "alice@example.com",
        "recipient": "bob@example.com",
        "action_items": ["Reset password", "Review account"],
        "is_spam": False,
    }
    input_jsonl.write_text(
        json.dumps(
            {
                "custom_id": "sample.eml",
                "response": {
                    "status_code": 200,
                    "body": {
                        "output": [
                            {
                                "type": "message",
                                "content": [
                                    {
                                        "type": "output_text",
                                        "text": json.dumps(
                                            output_payload,
                                            ensure_ascii=False,
                                            sort_keys=True,
                                        ),
                                    }
                                ],
                            }
                        ]
                    },
                },
            },
            ensure_ascii=False,
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )

    exit_code = run_batch_output_to_xlsx(
        BatchOutputXlsxConfig(
            input_jsonl=input_jsonl,
            output_xlsx=output_xlsx,
            schema_file=schema_file,
        )
    )

    assert exit_code == 0
    assert output_xlsx.exists()

    workbook = load_workbook(output_xlsx)
    worksheet = workbook.active

    assert [cell.value for cell in worksheet[1]] == [
        "filename",
        "from",
        "recipient",
        "action_items",
        "is_spam",
    ]
    assert [cell.value for cell in worksheet[2]] == [
        "sample.eml",
        "alice@example.com",
        "bob@example.com",
        '["Reset password", "Review account"]',
        False,
    ]


def test_batch_output_to_xlsx_infers_columns_without_schema_file(
    tmp_path: Path,
) -> None:
    input_jsonl = tmp_path / "batch_output.jsonl"
    output_xlsx = tmp_path / "batch_output.xlsx"

    input_jsonl.write_text(
        "\n".join(
            [
                json.dumps(
                    {
                        "custom_id": "first.eml",
                        "response": {
                            "status_code": 200,
                            "body": {
                                "output": [
                                    {
                                        "type": "message",
                                        "content": [
                                            {
                                                "type": "output_text",
                                                "text": '{"subject":"Hello","score":1.5}',
                                            }
                                        ],
                                    }
                                ]
                            },
                        },
                    },
                    sort_keys=True,
                ),
                json.dumps(
                    {
                        "custom_id": "second.eml",
                        "response": {
                            "status_code": 200,
                            "body": {
                                "output": [
                                    {
                                        "type": "message",
                                        "content": [
                                            {
                                                "type": "output_text",
                                                "text": '{"subject":"World","score":2.5,"category":"ham"}',
                                            }
                                        ],
                                    }
                                ]
                            },
                        },
                    },
                    sort_keys=True,
                ),
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    exit_code = run_batch_output_to_xlsx(
        BatchOutputXlsxConfig(
            input_jsonl=input_jsonl,
            output_xlsx=output_xlsx,
        )
    )

    assert exit_code == 0

    workbook = load_workbook(output_xlsx)
    worksheet = workbook.active

    assert [cell.value for cell in worksheet[1]] == [
        "filename",
        "subject",
        "score",
        "category",
    ]
    assert [cell.value for cell in worksheet[2]] == ["first.eml", "Hello", 1.5, None]
    assert [cell.value for cell in worksheet[3]] == ["second.eml", "World", 2.5, "ham"]
