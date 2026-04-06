"""Evaluate benchmark results against ground-truth labels.

Reads a ``batch_output.jsonl`` file (produced by ``submit-batch`` or
``submit-ollama-batch``), infers ground-truth labels from the ``custom_id``
filename prefix, and computes a confusion matrix with per-class
precision / recall / F1.  Results are printed to the terminal and
optionally written to an XLSX workbook.
"""

from __future__ import annotations

import json
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import EvalBenchmarkConfig

# openpyxl rejects characters that are illegal in XML 1.0.  We strip them
# before writing cell values so garbled email content doesn't crash the
# XLSX export.
_ILLEGAL_XML_CHARS_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f"
    r"\ud800-\udfff\ufffe\uffff]"
)


def _sanitise_for_xlsx(value: Any) -> Any:
    """Strip characters that are illegal in XML 1.0 from string values."""
    if isinstance(value, str):
        return _ILLEGAL_XML_CHARS_RE.sub("", value)
    return value


# ---------------------------------------------------------------------------
# Data containers
# ---------------------------------------------------------------------------


@dataclass(slots=True)
class PredictionRow:
    """One email's ground truth vs. model prediction."""

    custom_id: str
    ground_truth: str
    predicted: str
    confidence: float | None = None
    raw_output: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class ClassMetrics:
    """Precision / recall / F1 for a single class label."""

    label: str
    true_positives: int = 0
    false_positives: int = 0
    false_negatives: int = 0

    @property
    def precision(self) -> float:
        denom = self.true_positives + self.false_positives
        return self.true_positives / denom if denom > 0 else 0.0

    @property
    def recall(self) -> float:
        denom = self.true_positives + self.false_negatives
        return self.true_positives / denom if denom > 0 else 0.0

    @property
    def f1(self) -> float:
        p, r = self.precision, self.recall
        return 2 * p * r / (p + r) if (p + r) > 0 else 0.0

    @property
    def support(self) -> int:
        return self.true_positives + self.false_negatives


@dataclass(slots=True)
class EvalReport:
    """Full evaluation report."""

    rows: list[PredictionRow]
    labels: list[str]
    confusion: dict[str, dict[str, int]]  # confusion[truth][predicted] = count
    per_class: dict[str, ClassMetrics]
    accuracy: float
    macro_precision: float
    macro_recall: float
    macro_f1: float
    weighted_precision: float
    weighted_recall: float
    weighted_f1: float
    total: int
    correct: int


# ---------------------------------------------------------------------------
# Ground-truth extraction
# ---------------------------------------------------------------------------

DEFAULT_CATEGORY_MAP: dict[str, str] = {
    "easy_ham": "ham",
    "easy_ham_2": "ham",
    "hard_ham": "ham",
    "spam": "spam",
    "spam_2": "spam",
}


def _infer_ground_truth(
    custom_id: str,
    category_map: dict[str, str],
) -> str:
    """Derive the ground-truth label from a filename-style ``custom_id``.

    The convention is ``<category>__<sequence>.<hash>`` — e.g.
    ``easy_ham__00001.7c53336b37003a9286aba55d2945844c``.
    """
    # Split on the first double-underscore to extract the category prefix.
    if "__" in custom_id:
        prefix = custom_id.split("__", 1)[0]
    else:
        prefix = custom_id

    normalised = prefix.lower().strip()
    if normalised in category_map:
        return category_map[normalised]

    return normalised


# ---------------------------------------------------------------------------
# Batch output parsing
# ---------------------------------------------------------------------------


def _extract_output_text(response_body: dict[str, Any]) -> str | None:
    """Pull the first ``output_text`` fragment from an OpenAI-shaped response body."""
    output_items = response_body.get("output")
    if not isinstance(output_items, list):
        return None

    fragments: list[str] = []
    for output_item in output_items:
        if not isinstance(output_item, dict):
            continue
        content_items = output_item.get("content")
        if not isinstance(content_items, list):
            continue
        for content_item in content_items:
            if not isinstance(content_item, dict):
                continue
            if content_item.get("type") != "output_text":
                continue
            text = content_item.get("text")
            if isinstance(text, str):
                fragments.append(text)

    return "".join(fragments) if fragments else None


def _load_predictions(
    batch_output_jsonl: Path,
    label_field: str,
    category_map: dict[str, str],
) -> list[PredictionRow]:
    """Parse the batch output JSONL and build prediction rows."""
    rows: list[PredictionRow] = []

    with batch_output_jsonl.open("r", encoding="utf-8", newline="") as fh:
        for _line_number, raw_line in enumerate(fh, start=1):
            text = raw_line.rstrip("\r\n")
            if not text.strip():
                continue

            record = json.loads(text)
            custom_id = record.get("custom_id", "")

            # Skip error records
            if record.get("error") is not None:
                continue

            response = record.get("response", {})
            if not isinstance(response, dict):
                continue

            body = response.get("body", {})
            if not isinstance(body, dict):
                continue

            output_text = _extract_output_text(body)
            if output_text is None:
                continue

            try:
                parsed = json.loads(output_text)
            except (json.JSONDecodeError, TypeError):
                parsed = {}

            if not isinstance(parsed, dict):
                parsed = {}

            predicted_raw = parsed.get(label_field, "")
            predicted = str(predicted_raw).strip().lower() if predicted_raw else ""
            confidence_raw = parsed.get("confidence")
            confidence: float | None = None
            if isinstance(confidence_raw, (int, float)):
                confidence = float(confidence_raw)

            ground_truth = _infer_ground_truth(custom_id, category_map)

            rows.append(
                PredictionRow(
                    custom_id=custom_id,
                    ground_truth=ground_truth,
                    predicted=predicted,
                    confidence=confidence,
                    raw_output=parsed,
                )
            )

    return rows


# ---------------------------------------------------------------------------
# Metric computation
# ---------------------------------------------------------------------------


def _compute_report(rows: list[PredictionRow]) -> EvalReport:
    """Build an ``EvalReport`` from prediction rows."""
    # Discover all labels (ground truth and predicted, sorted).
    label_set: set[str] = set()
    for row in rows:
        label_set.add(row.ground_truth)
        label_set.add(row.predicted)
    label_set.discard("")
    labels = sorted(label_set)

    # Build confusion matrix.
    confusion: dict[str, dict[str, int]] = {
        gt: {pred: 0 for pred in labels} for gt in labels
    }
    for row in rows:
        gt = row.ground_truth
        pred = row.predicted
        if gt in confusion and pred in confusion[gt]:
            confusion[gt][pred] += 1

    # Per-class metrics.
    per_class: dict[str, ClassMetrics] = {}
    for label in labels:
        tp = confusion.get(label, {}).get(label, 0)
        fp = sum(
            confusion.get(other, {}).get(label, 0) for other in labels if other != label
        )
        fn = sum(
            confusion.get(label, {}).get(other, 0) for other in labels if other != label
        )
        per_class[label] = ClassMetrics(
            label=label,
            true_positives=tp,
            false_positives=fp,
            false_negatives=fn,
        )

    total = len(rows)
    correct = sum(1 for r in rows if r.ground_truth == r.predicted)
    accuracy = correct / total if total > 0 else 0.0

    # Macro averages.
    n_classes = len(labels)
    macro_precision = (
        sum(m.precision for m in per_class.values()) / n_classes
        if n_classes > 0
        else 0.0
    )
    macro_recall = (
        sum(m.recall for m in per_class.values()) / n_classes if n_classes > 0 else 0.0
    )
    macro_f1 = (
        sum(m.f1 for m in per_class.values()) / n_classes if n_classes > 0 else 0.0
    )

    # Weighted averages.
    total_support = sum(m.support for m in per_class.values())
    if total_support > 0:
        weighted_precision = (
            sum(m.precision * m.support for m in per_class.values()) / total_support
        )
        weighted_recall = (
            sum(m.recall * m.support for m in per_class.values()) / total_support
        )
        weighted_f1 = sum(m.f1 * m.support for m in per_class.values()) / total_support
    else:
        weighted_precision = 0.0
        weighted_recall = 0.0
        weighted_f1 = 0.0

    return EvalReport(
        rows=rows,
        labels=labels,
        confusion=confusion,
        per_class=per_class,
        accuracy=accuracy,
        macro_precision=macro_precision,
        macro_recall=macro_recall,
        macro_f1=macro_f1,
        weighted_precision=weighted_precision,
        weighted_recall=weighted_recall,
        weighted_f1=weighted_f1,
        total=total,
        correct=correct,
    )


# ---------------------------------------------------------------------------
# Terminal output
# ---------------------------------------------------------------------------


def _print_report(report: EvalReport) -> None:
    """Print the evaluation report to stdout."""
    print("\n=== Evaluation Report ===\n")
    print(f"Total emails:  {report.total}")
    print(f"Correct:       {report.correct}")
    print(f"Accuracy:      {report.accuracy:.4f}")
    print()

    # Confusion matrix.
    print("Confusion Matrix (rows=ground truth, cols=predicted):\n")
    col_width = max(len(label) for label in report.labels) + 2
    col_width = max(col_width, 8)
    header = " " * col_width + "".join(
        label.rjust(col_width) for label in report.labels
    )
    print(header)
    for gt_label in report.labels:
        row_cells = "".join(
            str(report.confusion[gt_label].get(pred_label, 0)).rjust(col_width)
            for pred_label in report.labels
        )
        print(f"{gt_label.rjust(col_width)}{row_cells}")
    print()

    # Per-class metrics.
    print(f"{'Class':<15} {'Precision':>10} {'Recall':>10} {'F1':>10} {'Support':>10}")
    print("-" * 55)
    for label in report.labels:
        m = report.per_class[label]
        print(
            f"{label:<15} {m.precision:>10.4f} {m.recall:>10.4f} "
            f"{m.f1:>10.4f} {m.support:>10d}"
        )
    print("-" * 55)
    print(
        f"{'macro avg':<15} {report.macro_precision:>10.4f} "
        f"{report.macro_recall:>10.4f} {report.macro_f1:>10.4f} {report.total:>10d}"
    )
    print(
        f"{'weighted avg':<15} {report.weighted_precision:>10.4f} "
        f"{report.weighted_recall:>10.4f} {report.weighted_f1:>10.4f} {report.total:>10d}"
    )
    print()


# ---------------------------------------------------------------------------
# XLSX report
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center")


def _write_xlsx(report: EvalReport, output_xlsx: Path) -> None:
    """Write an XLSX workbook with three sheets: Summary, Confusion, Raw Data."""
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_rows: list[tuple[str, Any]] = [
        ("Total Emails", report.total),
        ("Correct", report.correct),
        ("Accuracy", report.accuracy),
        ("", ""),
        ("Macro Precision", report.macro_precision),
        ("Macro Recall", report.macro_recall),
        ("Macro F1", report.macro_f1),
        ("", ""),
        ("Weighted Precision", report.weighted_precision),
        ("Weighted Recall", report.weighted_recall),
        ("Weighted F1", report.weighted_f1),
    ]
    for row in summary_rows:
        ws_summary.append(row)

    ws_summary.append(("", ""))
    ws_summary.append(("", ""))

    # Per-class table.
    class_header = ("Class", "Precision", "Recall", "F1", "Support")
    ws_summary.append(class_header)
    class_header_row = ws_summary.max_row
    for col_idx in range(1, len(class_header) + 1):
        cell = ws_summary.cell(row=class_header_row, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER

    for label in report.labels:
        m = report.per_class[label]
        ws_summary.append((label, m.precision, m.recall, m.f1, m.support))

    ws_summary.append(
        (
            "macro avg",
            report.macro_precision,
            report.macro_recall,
            report.macro_f1,
            report.total,
        )
    )
    ws_summary.append(
        (
            "weighted avg",
            report.weighted_precision,
            report.weighted_recall,
            report.weighted_f1,
            report.total,
        )
    )

    ws_summary.column_dimensions["A"].width = 22
    for col_letter in ("B", "C", "D", "E"):
        ws_summary.column_dimensions[col_letter].width = 14

    # --- Sheet 2: Confusion Matrix ---
    ws_confusion = wb.create_sheet("Confusion Matrix")

    ws_confusion.append([""] + report.labels)
    header_row = ws_confusion.max_row
    for col_idx in range(2, len(report.labels) + 2):
        cell = ws_confusion.cell(row=header_row, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER

    for gt_label in report.labels:
        row_data: list[Any] = [gt_label]
        for pred_label in report.labels:
            row_data.append(report.confusion[gt_label].get(pred_label, 0))
        ws_confusion.append(row_data)
        row_num = ws_confusion.max_row
        ws_confusion.cell(row=row_num, column=1).font = _BOLD_FONT

    for col_idx in range(1, len(report.labels) + 2):
        ws_confusion.column_dimensions[get_column_letter(col_idx)].width = 14

    # --- Sheet 3: Raw Data ---
    ws_raw = wb.create_sheet("Raw Data")

    raw_header = ["custom_id", "ground_truth", "predicted", "correct", "confidence"]
    # Add any extra fields from the first row's raw_output.
    extra_keys: list[str] = []
    if report.rows:
        extra_keys = sorted(
            k
            for k in report.rows[0].raw_output
            if k not in {"classification", "confidence"}
        )
        raw_header.extend(extra_keys)

    ws_raw.append(raw_header)
    raw_header_row = ws_raw.max_row
    for col_idx in range(1, len(raw_header) + 1):
        cell = ws_raw.cell(row=raw_header_row, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER

    for row in report.rows:
        is_correct = row.ground_truth == row.predicted
        values: list[Any] = [
            row.custom_id,
            row.ground_truth,
            row.predicted,
            is_correct,
            row.confidence,
        ]
        for k in extra_keys:
            val = row.raw_output.get(k)
            if isinstance(val, (list, dict)):
                values.append(json.dumps(val, ensure_ascii=False, sort_keys=True))
            else:
                values.append(val)
        ws_raw.append([_sanitise_for_xlsx(v) for v in values])

    ws_raw.freeze_panes = "A2"
    ws_raw.auto_filter.ref = ws_raw.dimensions

    # Auto-size a few key columns.
    ws_raw.column_dimensions["A"].width = 50
    for col_letter in ("B", "C", "D", "E"):
        ws_raw.column_dimensions[col_letter].width = 14

    wb.save(output_xlsx)


# ---------------------------------------------------------------------------
# JSON report
# ---------------------------------------------------------------------------


def _write_json_report(report: EvalReport, output_path: Path) -> None:
    """Write a machine-readable JSON summary alongside the XLSX."""
    data: dict[str, Any] = {
        "total": report.total,
        "correct": report.correct,
        "accuracy": report.accuracy,
        "macro_precision": report.macro_precision,
        "macro_recall": report.macro_recall,
        "macro_f1": report.macro_f1,
        "weighted_precision": report.weighted_precision,
        "weighted_recall": report.weighted_recall,
        "weighted_f1": report.weighted_f1,
        "labels": report.labels,
        "per_class": {
            label: {
                "precision": m.precision,
                "recall": m.recall,
                "f1": m.f1,
                "support": m.support,
                "true_positives": m.true_positives,
                "false_positives": m.false_positives,
                "false_negatives": m.false_negatives,
            }
            for label, m in report.per_class.items()
        },
        "confusion_matrix": report.confusion,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(data, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def run_eval_benchmark(config: EvalBenchmarkConfig) -> int:
    """Run the evaluation benchmark and return an exit code."""
    # Load category map.
    category_map: dict[str, str]
    if config.category_map_file is not None:
        raw = json.loads(config.category_map_file.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            raise ValueError(
                f"Category map must be a JSON object: {config.category_map_file}"
            )
        category_map = {str(k): str(v) for k, v in raw.items()}
    else:
        category_map = dict(DEFAULT_CATEGORY_MAP)

    # Load predictions.
    rows = _load_predictions(
        config.batch_output_jsonl,
        label_field=config.label_field,
        category_map=category_map,
    )

    if not rows:
        print(f"No valid prediction rows found in {config.batch_output_jsonl}")
        return 1

    # Ground-truth distribution.
    gt_dist = Counter(r.ground_truth for r in rows)
    print(f"Loaded {len(rows)} predictions. Ground-truth distribution: {dict(gt_dist)}")

    # Compute metrics.
    report = _compute_report(rows)

    # Print to terminal.
    _print_report(report)

    # Write XLSX if requested.
    output_xlsx = config.output_xlsx
    if output_xlsx is None:
        output_xlsx = config.batch_output_jsonl.with_suffix(".eval.xlsx")

    _write_xlsx(report, output_xlsx)
    print(f"XLSX report written to: {output_xlsx}")

    # Also write a JSON report alongside.
    json_path = output_xlsx.with_suffix(".json")
    _write_json_report(report, json_path)
    print(f"JSON report written to: {json_path}")

    return 0
